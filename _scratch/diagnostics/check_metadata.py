import openpyxl

# Check baseline file
wb = openpyxl.load_workbook('baseline_data/baseline_recording_20251219_182111.xlsx')
print('Sheets:', wb.sheetnames)
print()

if 'Metadata' in wb.sheetnames:
    ws = wb['Metadata']
    print('=== METADATA ===')
    for row in ws.iter_rows(min_row=1):
        if row[0].value:
            print(f'{row[0].value}: {row[1].value if len(row) > 1 else ""}')
