"""Excel version manager for Affilabs.core experiment files.

Provides in-place versioning and audit history for .xlsx files produced by the
live recorder and subsequently edited in the Edits tab.

Contract
--------
- **Raw Data** sheet is write-protected on first creation (no password — structural
  lock to prevent accidental edits).  Protection is re-applied on every save.
- **Cycles** (and optionally Per-Channel Format) sheets are versioned: the active
  sheet is preserved and the previous version is renamed to ``Cycles_vN`` and hidden.
- **Edit History** sheet is an append-only audit log: one row per save event with
  timestamp, user, action description, and which cycles were modified.
- All version management happens *after* pandas has finished writing the new workbook,
  using direct openpyxl manipulation so we don't fight pandas' ExcelWriter.

Usage
-----
Call ``ExcelVersionManager.apply(file_path, action, user, cycles_affected, notes)``
after the new .xlsx has been written to disk.  The method opens the workbook,
applies protection, rotates versioned sheets, and appends a history row — all in
one atomic save.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    pass


# Sheets that are versioned (snapshot kept on every Edits save)
_VERSIONED_SHEETS = ("Cycles",)

# Sheet that is write-protected on every save
_PROTECTED_SHEET = "Raw Data"

# Name of the audit-log sheet
_HISTORY_SHEET = "Edit History"

# openpyxl sheet state values
_STATE_HIDDEN = "hidden"
_STATE_VISIBLE = "visible"


class ExcelVersionManager:
    """Apply version management to an existing .xlsx file in-place."""

    @staticmethod
    def apply(
        file_path: str | Path,
        *,
        action: str,
        user: str,
        cycles_affected: str = "",
        fields_changed: str = "",
        notes: str = "",
    ) -> None:
        """Open ``file_path``, apply versioning + protection, and save in-place.

        Parameters
        ----------
        file_path:
            Path to the .xlsx file that was just written by pandas.
        action:
            Human-readable description of what changed, e.g. "Exported selection".
        user:
            Name of the user who made the change.
        cycles_affected:
            Comma-separated list of cycle IDs or numbers that were touched.
        fields_changed:
            Comma-separated list of column names that were modified.
        notes:
            Any extra context (alignment shift applied, smoothing level, etc.).
        """
        try:
            import openpyxl
            from affilabs.utils.logger import logger
        except ImportError:
            return  # openpyxl not available — skip silently

        file_path = Path(file_path)
        if not file_path.exists():
            return

        try:
            wb = openpyxl.load_workbook(file_path)

            # 1. Rotate versioned sheets -----------------------------------------
            for sheet_name in _VERSIONED_SHEETS:
                if sheet_name in wb.sheetnames:
                    ExcelVersionManager._rotate_sheet(wb, sheet_name)

            # 2. Lock Raw Data sheet ------------------------------------------------
            if _PROTECTED_SHEET in wb.sheetnames:
                ws_raw = wb[_PROTECTED_SHEET]
                ws_raw.protection.sheet = True
                ws_raw.protection.formatColumns = False
                ws_raw.protection.formatRows = False
                ws_raw.protection.selectLockedCells = False
                ws_raw.protection.selectUnlockedCells = False
                # Do NOT set a password — structural lock only

            # 3. Append to Edit History --------------------------------------------
            ExcelVersionManager._append_history(
                wb,
                timestamp=datetime.now().isoformat(timespec="seconds"),
                user=user,
                action=action,
                cycles_affected=cycles_affected,
                fields_changed=fields_changed,
                notes=notes,
            )

            # 4. Move Edit History to end so it's the last tab --------------------
            if _HISTORY_SHEET in wb.sheetnames:
                wb.move_sheet(_HISTORY_SHEET, offset=len(wb.sheetnames))

            wb.save(file_path)
            logger.info(
                f"[VersionManager] Applied versioning to {file_path.name} "
                f"(user={user}, action={action!r})"
            )

        except Exception as e:
            try:
                from affilabs.utils.logger import logger
                logger.warning(f"[VersionManager] Failed to apply versioning to {file_path.name}: {e}")
            except Exception:
                pass

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _rotate_sheet(wb, sheet_name: str) -> None:
        """Rename the active sheet to the next versioned name and hide it.

        Only rotates if there is already a previous version (``Cycles_v1``
        etc.).  On the very first save the active sheet must remain as-is so
        the loader can find it by its canonical name.
        """
        existing = [s for s in wb.sheetnames if s.startswith(f"{sheet_name}_v")]
        if not existing:
            # First save — nothing to archive yet; keep the active sheet.
            return

        next_n = len(existing) + 1
        ws = wb[sheet_name]
        ws.title = f"{sheet_name}_v{next_n}"
        ws.sheet_state = _STATE_HIDDEN

    @staticmethod
    def _append_history(
        wb,
        *,
        timestamp: str,
        user: str,
        action: str,
        cycles_affected: str,
        fields_changed: str,
        notes: str,
    ) -> None:
        """Append a single row to the Edit History sheet (create if absent)."""
        headers = [
            "Timestamp", "User", "Action",
            "Cycles Affected", "Fields Changed", "Notes",
        ]

        if _HISTORY_SHEET not in wb.sheetnames:
            ws = wb.create_sheet(_HISTORY_SHEET)
            ws.append(headers)
            # Bold the header row
            try:
                from openpyxl.styles import Font
                for cell in ws[1]:
                    cell.font = Font(bold=True)
            except Exception:
                pass
            # Freeze header row
            ws.freeze_panes = "A2"
            # Column widths
            widths = [22, 14, 30, 20, 30, 40]
            for col_idx, width in enumerate(widths, 1):
                ws.column_dimensions[
                    ws.cell(row=1, column=col_idx).column_letter
                ].width = width
        else:
            ws = wb[_HISTORY_SHEET]
            # Make sure it's visible
            ws.sheet_state = _STATE_VISIBLE

        ws.append([timestamp, user, action, cycles_affected, fields_changed, notes])
