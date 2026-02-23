"""Excel Chart Builder for Post-Edit Analysis Export.

Creates interactive charts within Excel workbooks using openpyxl.
Charts include:
- Delta SPR bar charts
- Time-series line charts for cycle segments
- Flag position timeline charts
- Complete experiment overview charts
"""

from typing import List, Dict, Any, Optional
import pandas as pd
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, ScatterChart, Reference
    from openpyxl.chart.axis import DateAxis, ValuesAxis 
    from openpyxl.chart.data_source import NumDataSource, StrDataSource
    from openpyxl.chart.series import Series
    from openpyxl.styles import Font, NamedStyle
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExcelChartBuilder:
    """Builds interactive Excel charts for post-edit analysis exports."""
    
    def __init__(self, workbook: 'Workbook'):
        """Initialize chart builder with existing workbook.
        
        Args:
            workbook: openpyxl Workbook instance where charts will be added
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for chart creation")
            
        self.workbook = workbook
        
    def add_delta_spr_charts(self, analysis_data: pd.DataFrame) -> None:
        """Add Delta SPR bar charts showing channel responses.
        
        Creates one chart per cycle showing A/B/C/D channel delta values.
        
        Args:
            analysis_data: DataFrame with columns ['Cycle_ID', 'Delta_SPR_A', 'Delta_SPR_B', 'Delta_SPR_C', 'Delta_SPR_D']
        """
        if analysis_data.empty:
            return
            
        # Create Charts_Delta_SPR sheet
        chart_sheet = self.workbook.create_sheet("Charts_Delta_SPR")
        
        # Write analysis data to sheet for chart reference
        for row_idx, row_data in enumerate(dataframe_to_rows(analysis_data, index=False, header=True)):
            for col_idx, value in enumerate(row_data, 1):
                chart_sheet.cell(row=row_idx + 1, column=col_idx, value=value)
        
        # Create bar chart for each cycle
        row_offset = len(analysis_data) + 3  # Start charts below data
        
        for idx, (_, cycle_row) in enumerate(analysis_data.iterrows()):
            cycle_id = cycle_row.get('Cycle_ID', f'Cycle_{idx+1}')
            
            # Create bar chart
            bar_chart = BarChart()
            bar_chart.type = "col"
            bar_chart.style = 10
            bar_chart.title = f"Delta SPR - {cycle_id}"
            bar_chart.y_axis.title = "Response (RU)"
            bar_chart.x_axis.title = "Channel"
            
            # Data for channels A, B, C, D (columns 2-5 in analysis_data)
            data_ref = Reference(chart_sheet, 
                               min_col=2, 
                               max_col=5, 
                               min_row=idx + 2, 
                               max_row=idx + 2)
            
            # Channel labels
            labels_ref = Reference(chart_sheet, 
                                 min_col=1, 
                                 max_col=1, 
                                 min_row=1, 
                                 max_row=1)
            
            # Manually set channel labels
            chart_sheet.cell(row=len(analysis_data) + 5, column=1, value="Channel A")
            chart_sheet.cell(row=len(analysis_data) + 5, column=2, value="Channel B") 
            chart_sheet.cell(row=len(analysis_data) + 5, column=3, value="Channel C")
            chart_sheet.cell(row=len(analysis_data) + 5, column=4, value="Channel D")
            
            # Add data series
            bar_chart.add_data(data_ref, titles_from_data=False)
            
            # Set channel labels manually
            bar_chart.x_axis.title = "Channel"
            
            # Position chart
            chart_position = f"A{row_offset + (idx * 15)}"
            chart_sheet.add_chart(bar_chart, chart_position)
            
    def add_timeline_charts(self, processed_data: pd.DataFrame, cycles_data: pd.DataFrame) -> None:
        """Add time-series line charts for cycle segments.
        
        Creates line charts showing SPR response over time for each cycle.
        
        Args:
            processed_data: DataFrame with time and SPR columns for each channel
            cycles_data: DataFrame with cycle start/stop times
        """
        if processed_data.empty or cycles_data.empty:
            return
            
        # Create Charts_Timeline sheet
        chart_sheet = self.workbook.create_sheet("Charts_Timeline")
        
        # Write processed data to sheet
        for row_idx, row_data in enumerate(dataframe_to_rows(processed_data, index=False, header=True)):
            for col_idx, value in enumerate(row_data, 1):
                chart_sheet.cell(row=row_idx + 1, column=col_idx, value=value)
        
        data_rows = len(processed_data) + 1
        
        # Create timeline chart for each cycle
        for idx, (_, cycle_row) in enumerate(cycles_data.iterrows()):
            cycle_id = cycle_row.get('cycle_id', f'Cycle_{idx+1}')
            start_time = cycle_row.get('start_time_sensorgram', 0)
            end_time = cycle_row.get('end_time_sensorgram', 100)
            
            # Create line chart
            line_chart = LineChart()
            line_chart.title = f"Timeline - {cycle_id}"
            line_chart.style = 2
            line_chart.x_axis.title = "Time (s)"
            line_chart.y_axis.title = "SPR Response (RU)"
            
            # Assume Time column is column 1, SPR columns are 2, 4, 6, 8 (for A, B, C, D)
            time_ref = Reference(chart_sheet, min_col=1, max_col=1, min_row=2, max_row=data_rows)
            
            # Add each channel as a series
            channel_colors = ['FF0000', '0000FF', '00FF00', 'FF8000']  # Red, Blue, Green, Orange
            for ch_idx, (ch_name, color) in enumerate(zip(['A', 'B', 'C', 'D'], channel_colors)):
                spr_col = 2 + (ch_idx * 2) + 1  # Assuming Time_A, SPR_A, Time_B, SPR_B pattern
                if spr_col <= processed_data.shape[1]:
                    spr_ref = Reference(chart_sheet, min_col=spr_col, max_col=spr_col, min_row=2, max_row=data_rows)
                    series = Series(spr_ref, time_ref, title=f"Channel {ch_name}")
                    line_chart.append(series)
            
            # Position chart
            chart_position = f"J{2 + (idx * 20)}"
            chart_sheet.add_chart(line_chart, chart_position)
            
    def add_flags_timeline_chart(self, flag_data: pd.DataFrame, cycles_data: pd.DataFrame) -> None:
        """Add visual timeline chart showing flag positions.
        
        Creates scatter chart showing injection/wash/spike positions across cycles.
        
        Args:
            flag_data: DataFrame with flag positions and types
            cycles_data: DataFrame with cycle information
        """
        if flag_data.empty:
            return
            
        # Create Charts_Flags sheet
        chart_sheet = self.workbook.create_sheet("Charts_Flags")
        
        # Map flag types to numeric Y values for scatter chart (openpyxl requires numeric axes)
        flag_types = ['injection', 'wash', 'spike']
        type_y_value = {'injection': 1, 'wash': 2, 'spike': 3}
        type_colors = {'injection': 'FF0000', 'wash': '0000FF', 'spike': '00FF00'}

        # Write a legend table so user knows which Y value = which flag type
        chart_sheet.append(["Flag Type", "Y Value"])
        for ft in flag_types:
            chart_sheet.append([ft.title(), type_y_value[ft]])
        chart_sheet.append([])  # blank spacer row

        # Track row ranges per flag type for Reference objects
        # Data starts at row 5 (rows 1-3 = legend, row 4 = spacer)
        series_ranges = {}  # flag_type -> (start_row, end_row)
        current_row = 5

        for flag_type in flag_types:
            type_flags = (
                flag_data[flag_data['Flag_Type'] == flag_type]
                if 'Flag_Type' in flag_data.columns
                else pd.DataFrame()
            )
            if type_flags.empty:
                continue

            # Write header
            chart_sheet.append(["Time (s)", "Y", "Cycle"])
            current_row += 1
            block_start = current_row

            for _, flag_row in type_flags.iterrows():
                time_pos = flag_row.get('Time_Position', 0)
                cycle_id = flag_row.get('Cycle_ID', 'Unknown')
                chart_sheet.append([time_pos, type_y_value[flag_type], str(cycle_id)])
                current_row += 1

            series_ranges[flag_type] = (block_start, current_row - 1)
            chart_sheet.append([])  # blank spacer between blocks
            current_row += 1

        if not series_ranges:
            return

        # Create scatter chart with numeric Y axis
        scatter_chart = ScatterChart()
        scatter_chart.title = "Flag Positions Timeline"
        scatter_chart.style = 2
        scatter_chart.x_axis.title = "Time (s)"
        scatter_chart.y_axis.title = "Flag Type (1=Injection, 2=Wash, 3=Spike)"
        scatter_chart.y_axis.numFmt = '0'
        scatter_chart.y_axis.scaling.min = 0
        scatter_chart.y_axis.scaling.max = 4
        scatter_chart.width = 20
        scatter_chart.height = 10

        for flag_type, (start_row, end_row) in series_ranges.items():
            x_values = Reference(chart_sheet, min_col=1, min_row=start_row, max_row=end_row)
            y_values = Reference(chart_sheet, min_col=2, min_row=start_row, max_row=end_row)
            series = Series(y_values, x_values, title=flag_type.title())
            series.marker.symbol = "circle"
            series.marker.size = 8
            series.graphicalProperties.line.noFill = True
            series.graphicalProperties.solidFill = type_colors[flag_type]
            scatter_chart.series.append(series)

        chart_sheet.add_chart(scatter_chart, "F2")
        
    def add_overview_chart(self, processed_data: pd.DataFrame, cycles_data: pd.DataFrame) -> None:
        """Add complete experiment overview chart.
        
        Shows all cycles on one timeline with cycle boundaries marked.
        
        Args:
            processed_data: Complete time-series data
            cycles_data: Cycle boundary information
        """
        # Create Charts_Overview sheet
        chart_sheet = self.workbook.create_sheet("Charts_Overview")
        
        # Write processed data
        for row_idx, row_data in enumerate(dataframe_to_rows(processed_data, index=False, header=True)):
            for col_idx, value in enumerate(row_data, 1):
                chart_sheet.cell(row=row_idx + 1, column=col_idx, value=value)
        
        data_rows = len(processed_data) + 1
        
        # Create overview line chart
        overview_chart = LineChart()
        overview_chart.title = "Complete Experiment Overview"
        overview_chart.style = 1
        overview_chart.width = 20
        overview_chart.height = 10
        overview_chart.x_axis.title = "Time (s)"
        overview_chart.y_axis.title = "SPR Response (RU)"
        
        # Time reference
        time_ref = Reference(chart_sheet, min_col=1, max_col=1, min_row=2, max_row=data_rows)
        
        # Add all channels
        channel_names = ['A', 'B', 'C', 'D']
        for ch_idx, ch_name in enumerate(channel_names):
            spr_col = 2 + (ch_idx * 2) + 1  # SPR column for each channel
            if spr_col <= processed_data.shape[1]:
                spr_ref = Reference(chart_sheet, min_col=spr_col, max_col=spr_col, min_row=2, max_row=data_rows)
                series = Series(spr_ref, time_ref, title=f"Channel {ch_name}")
                overview_chart.append(series)
        
        # Position the large overview chart
        chart_sheet.add_chart(overview_chart, "F2")
        
        # Add cycle boundary markers as text annotations
        cycles_info_start_row = data_rows + 5
        chart_sheet.cell(row=cycles_info_start_row, column=1, value="Cycle Boundaries")
        chart_sheet.cell(row=cycles_info_start_row + 1, column=1, value="Cycle ID")
        chart_sheet.cell(row=cycles_info_start_row + 1, column=2, value="Start (s)")
        chart_sheet.cell(row=cycles_info_start_row + 1, column=3, value="End (s)")
        
        for idx, (_, cycle_row) in enumerate(cycles_data.iterrows()):
            row = cycles_info_start_row + 2 + idx
            chart_sheet.cell(row=row, column=1, value=cycle_row.get('cycle_id', f'Cycle_{idx+1}'))
            chart_sheet.cell(row=row, column=2, value=cycle_row.get('start_time_sensorgram', 0))
            chart_sheet.cell(row=row, column=3, value=cycle_row.get('end_time_sensorgram', 100))

    def add_binding_plot_chart(self, binding_fit: dict) -> None:
        """Add Binding Plot sheet with data table and XY scatter chart.

        Args:
            binding_fit: Dict from BindingPlotMixin._binding_fit_result containing
                         'model', 'channel', 'r2', 'ref', 'conc', 'dspr', 'labels',
                         'x_fit', 'y_fit', 'params', and optionally 'Kd_uM', 'Rmax_RU'.
        """
        if not OPENPYXL_AVAILABLE:
            return

        from openpyxl.chart import ScatterChart, Reference, Series as ChartSeries

        sheet = self.workbook.create_sheet("Binding Plot")

        model   = binding_fit.get('model', 'Unknown')
        channel = binding_fit.get('channel', '?')
        r2      = binding_fit.get('r2', 0.0)
        ref     = binding_fit.get('ref', 'None')
        params  = binding_fit.get('params', '')
        conc    = binding_fit.get('conc', [])
        dspr    = binding_fit.get('dspr', [])
        labels  = binding_fit.get('labels', [])
        x_fit   = binding_fit.get('x_fit', [])
        y_fit   = binding_fit.get('y_fit', [])

        # --- Header ---
        sheet.cell(row=1, column=1, value=f"Binding Plot — Ch {channel} — {model}")
        sheet.cell(row=2, column=1, value=f"R² = {r2:.3f}   |   ref: {ref}   |   {params.replace(chr(10), '  ')}")

        # --- Raw data table (cols A–C starting row 4) ---
        sheet.cell(row=4, column=1, value="Cycle")
        sheet.cell(row=4, column=2, value="Concentration_uM")
        sheet.cell(row=4, column=3, value="Delta_SPR_RU")
        for i, (lbl, c, d) in enumerate(zip(labels, conc, dspr)):
            r = 5 + i
            sheet.cell(row=r, column=1, value=lbl)
            sheet.cell(row=r, column=2, value=float(c))
            sheet.cell(row=r, column=3, value=float(d))

        n_raw = len(conc)
        raw_last_row = 4 + n_raw  # inclusive

        # --- Fit line table (cols E–F starting row 4) ---
        sheet.cell(row=4, column=5, value="Fit_X_uM")
        sheet.cell(row=4, column=6, value="Fit_Y_RU")
        for i, (x, y) in enumerate(zip(x_fit, y_fit)):
            r = 5 + i
            sheet.cell(row=r, column=5, value=float(x))
            sheet.cell(row=r, column=6, value=float(y))

        n_fit = len(x_fit)
        fit_last_row = 4 + n_fit  # inclusive

        # --- Metadata (col H) ---
        sheet.cell(row=4, column=8, value="Channel")
        sheet.cell(row=5, column=8, value=channel)
        sheet.cell(row=6, column=8, value="Model")
        sheet.cell(row=7, column=8, value=model)
        sheet.cell(row=8, column=8, value="R2")
        sheet.cell(row=9, column=8, value=r2)
        sheet.cell(row=10, column=8, value="Reference")
        sheet.cell(row=11, column=8, value=ref)
        if 'Kd_uM' in binding_fit:
            sheet.cell(row=12, column=8, value="Kd_uM")
            sheet.cell(row=13, column=8, value=binding_fit['Kd_uM'])
        if 'Rmax_RU' in binding_fit:
            sheet.cell(row=14, column=8, value="Rmax_RU (empirical)")
            sheet.cell(row=15, column=8, value=binding_fit['Rmax_RU'])

        # --- Rmax Calculator summary block (col H, rows 17+) ---
        rmax_rows = [
            ("Ligand MW (Da)",        binding_fit.get('rmax_ligand_mw')),
            ("Analyte MW (Da)",       binding_fit.get('rmax_analyte_mw')),
            ("Immob ΔSPR (RU)",       binding_fit.get('rmax_immob_dspr_ru')),
            ("Theoretical Rmax (RU)", binding_fit.get('rmax_theoretical_ru')),
            ("Empirical Rmax (RU)",   binding_fit.get('Rmax_RU')),
            ("Surface activity (%)",  binding_fit.get('rmax_surface_activity_pct')),
        ]
        r_offset = 17
        sheet.cell(row=r_offset, column=8, value="Rmax Calculator")
        for i, (label, value) in enumerate(rmax_rows):
            sheet.cell(row=r_offset + 1 + i * 2, column=8, value=label)
            sheet.cell(row=r_offset + 2 + i * 2, column=8,
                       value=round(value, 2) if value is not None else "—")

        # --- XY Scatter chart ---
        if n_raw >= 2 and n_fit >= 2:
            chart = ScatterChart()
            chart.title = f"Binding Plot — Ch {channel} — {model}"
            chart.style = 10
            chart.x_axis.title = "Concentration (µM)"
            chart.y_axis.title = "ΔSPR (RU)"
            chart.width = 18
            chart.height = 12

            # Series 1: raw data points (markers only)
            x_ref_raw = Reference(sheet, min_col=2, min_row=5, max_row=raw_last_row)
            y_ref_raw = Reference(sheet, min_col=3, min_row=5, max_row=raw_last_row)
            ser_raw = ChartSeries(y_ref_raw, x_ref_raw, title=f"Ch {channel} data")
            ser_raw.graphicalProperties.line.noFill = True
            ser_raw.marker.symbol = "circle"
            ser_raw.marker.size = 7
            chart.series.append(ser_raw)

            # Series 2: fit line (line only, no markers)
            x_ref_fit = Reference(sheet, min_col=5, min_row=5, max_row=fit_last_row)
            y_ref_fit = Reference(sheet, min_col=6, min_row=5, max_row=fit_last_row)
            ser_fit = ChartSeries(y_ref_fit, x_ref_fit, title=f"{model} fit")
            ser_fit.marker.symbol = "none"
            chart.series.append(ser_fit)

            sheet.add_chart(chart, "A17")


def create_analysis_workbook_with_charts(
    raw_data: pd.DataFrame,
    processed_data: pd.DataFrame,
    analysis_results: pd.DataFrame,
    flag_data: pd.DataFrame,
    cycles_data: pd.DataFrame,
    export_settings: Dict[str, Any],
    output_path: Path,
    selected_cycles: list = None,
    binding_fit: dict = None,
) -> None:
    """Create complete analysis workbook with data sheets and interactive charts.

    Args:
        raw_data: Original untouched XY data
        processed_data: Post-edit processed curves
        analysis_results: Delta measurements and cursor positions
        flag_data: Updated marker positions
        cycles_data: Enhanced cycle metadata
        export_settings: Documentation of processing applied
        output_path: Where to save the Excel file
        selected_cycles: List of cycle indices to include in export (None = all)
        binding_fit: Optional dict from BindingPlotMixin._binding_fit_result
    """
    if not OPENPYXL_AVAILABLE:
        print("Warning: openpyxl not available. Creating basic Excel file without charts.")
        # Fallback to pandas ExcelWriter
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            raw_data.to_excel(writer, sheet_name='Raw_Data', index=False)
            processed_data.to_excel(writer, sheet_name='Processed_Data', index=False)
            analysis_results.to_excel(writer, sheet_name='Analysis_Results', index=False)
            flag_data.to_excel(writer, sheet_name='Flag_Positions', index=False)
            cycles_data.to_excel(writer, sheet_name='Cycles_Metadata', index=False)
            pd.DataFrame([export_settings]).to_excel(writer, sheet_name='Export_Settings', index=False)
        return
    
    # Create workbook with data sheets
    workbook = Workbook()
    
    # Remove default sheet
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    
    # Add data sheets
    sheets_data = {
        'Raw_Data': raw_data,
        'Processed_Data': processed_data, 
        'Analysis_Results': analysis_results,
        'Flag_Positions': flag_data,
        'Cycles_Metadata': cycles_data,
        'Export_Settings': pd.DataFrame([export_settings])
    }
    
    for sheet_name, df in sheets_data.items():
        if not df.empty:
            sheet = workbook.create_sheet(sheet_name)
            for row_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=True)):
                for col_idx, value in enumerate(row_data, 1):
                    sheet.cell(row=row_idx + 1, column=col_idx, value=value)
    
    # Add interactive charts
    try:
        chart_builder = ExcelChartBuilder(workbook)
        
        if not analysis_results.empty:
            chart_builder.add_delta_spr_charts(analysis_results)
            
        if not processed_data.empty and not cycles_data.empty:
            chart_builder.add_timeline_charts(processed_data, cycles_data)
            chart_builder.add_overview_chart(processed_data, cycles_data)
            
        if not flag_data.empty:
            chart_builder.add_flags_timeline_chart(flag_data, cycles_data)

        if binding_fit is not None:
            chart_builder.add_binding_plot_chart(binding_fit)

        print("✓ Added interactive Excel charts")
        
    except Exception as e:
        print(f"Warning: Could not add charts to Excel file: {e}")
    
    # Save workbook
    workbook.save(output_path)
    print(f"✓ Saved analysis workbook with charts: {output_path}")