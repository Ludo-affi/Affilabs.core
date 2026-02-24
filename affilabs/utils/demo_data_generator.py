"""Demo Data Generator for promotional UI screenshots.

Generates realistic SPR binding curves in wavelength domain (nm):
  - Live sensorgram: 5-concentration series (~30 min, all cycles visible)
  - Active cycle graph: last (highest) concentration cycle — full 360 s view
  - Edits tab: kinetics_demo.xlsx with matching concentration data

Ctrl+Shift+D: generates a full .xlsx demo file (saved to _data/demo/kinetics_demo.xlsx)
and loads it into the Edits tab via the normal Excel import path, then populates
the live graphs, Sparq, spectroscopy, and bar chart.

Call load_demo_data_into_app(app) from the main Application instance.
"""

import os
import time

import numpy as np


# Baseline SPR wavelength for gold sensors (~615 nm typical)
_BASELINE_NM = 615.0

# 4 distinct channel profiles: (rmax_nm, ka_rel, kd_rel, noise_nm)
# Different Rmax, association rate, dissociation rate per channel.
_CHANNEL_PROFILES = {
    "a": dict(rmax=0.9,  ka=0.4,  kd=3.5,  noise=0.004),  # Slow on, very fast off — low response
    "b": dict(rmax=1.6,  ka=0.6,  kd=2.5,  noise=0.004),  # Slower on, fast off — moderate
    "c": dict(rmax=2.1,  ka=1.4,  kd=1.2,  noise=0.004),  # Fast on, medium off
    "d": dict(rmax=2.8,  ka=1.0,  kd=0.8,  noise=0.004),  # Fast on, slow off — high response
}


def generate_spr_wavelength_curve(
    times: np.ndarray,
    baseline_nm: float,
    rmax_nm: float,
    ka_rel: float,
    kd_rel: float,
    noise_nm: float,
    assoc_start: float,
    assoc_end: float,
    dissoc_end: float,
    rng: np.random.Generator,
) -> np.ndarray:
    """Generate one SPR channel as a wavelength (nm) time series.

    SPR convention for this system: binding causes a RED SHIFT (wavelength increases).
    Signal starts at baseline_nm, rises during association, partially recovers during dissociation.
    """
    n = len(times)
    signal = np.full(n, baseline_nm, dtype=float)

    # Association phase
    assoc_mask = (times >= assoc_start) & (times < assoc_end)
    t_a = times[assoc_mask] - assoc_start
    kobs = ka_rel * 0.012  # scaled to give reasonable approach to Rmax over ~3 min
    response_assoc = rmax_nm * (1 - np.exp(-kobs * t_a))
    signal[assoc_mask] += response_assoc

    # Hold at end-of-association level
    if assoc_mask.any():
        assoc_end_response = response_assoc[-1]
    else:
        assoc_end_response = 0.0

    # Dissociation phase
    dissoc_mask = (times >= assoc_end) & (times < dissoc_end)
    t_d = times[dissoc_mask] - assoc_end
    kd_scaled = kd_rel * 0.008
    response_dissoc = assoc_end_response * np.exp(-kd_scaled * t_d)
    signal[dissoc_mask] += response_dissoc

    # After dissoc_end: residual offset (some channels don't fully return)
    if dissoc_mask.any():
        residual = response_dissoc[-1]
    else:
        residual = 0.0
    post_mask = times >= dissoc_end
    signal[post_mask] += residual

    signal += rng.normal(0, noise_nm, n)

    return signal


def generate_demo_sensorgrams(
    duration_s: float = 900.0,   # 15 minutes
    rate_hz: float = 2.0,
    assoc_start: float = 120.0,  # 2 min baseline
    assoc_duration: float = 300.0,  # 5 min association
    dissoc_duration: float = 360.0,  # 6 min dissociation (remaining)
    seed: int = 7,
) -> tuple[np.ndarray, dict[str, np.ndarray]]:
    """Generate 4-channel 15-min SPR wavelength sensorgrams.

    Returns:
        times: array of elapsed seconds (shape N)
        wl: dict {'a','b','c','d'} of wavelength arrays (nm, shape N)
    """
    rng = np.random.default_rng(seed)
    times = np.arange(0, duration_s, 1.0 / rate_hz)
    assoc_end = assoc_start + assoc_duration
    dissoc_end = assoc_end + dissoc_duration

    wl: dict[str, np.ndarray] = {}
    for ch, p in _CHANNEL_PROFILES.items():
        wl[ch] = generate_spr_wavelength_curve(
            times,
            baseline_nm=_BASELINE_NM,
            rmax_nm=p["rmax"],
            ka_rel=p["ka"],
            kd_rel=p["kd"],
            noise_nm=p["noise"],
            assoc_start=assoc_start,
            assoc_end=assoc_end,
            dissoc_end=dissoc_end,
            rng=rng,
        )

    return times, wl


def generate_concentration_series(
    concentrations_nm: list[float] = None,
    baseline_s: float = 60.0,    # baseline before each injection
    assoc_s: float = 120.0,      # association phase — short enough that curves don't all plateau
    dissoc_s: float = 180.0,     # dissociation phase
    rate_hz: float = 1.0,        # 1 Hz for Excel (smaller file)
    ka_per_Ms: float = 5e4,      # on-rate constant (M⁻¹s⁻¹) — slower so kobs varies with [C]
    kd_per_s: float = 2.5e-4,    # off-rate constant (s⁻¹) — overridden below to KD=150 nM
    seed: int = 42,
) -> tuple[np.ndarray, dict[str, dict[str, np.ndarray]]]:
    """Generate a multi-cycle concentration series for the Edits/Excel demo.

    Each concentration gets its own time segment:
      [baseline | association | dissociation] repeated N times

    The on-rate (kobs = ka*C + kd) scales with concentration so higher
    concentrations associate faster AND reach a higher plateau — exactly
    what a real dose-response experiment looks like.

    Returns:
        times_abs: flat absolute time array (all cycles concatenated)
        wl_by_cycle: dict {cycle_idx: {'a': array, 'b': array, ...}}
        cycle_windows: list of (t_start, t_end) for each cycle
    """
    if concentrations_nm is None:
        concentrations_nm = _DEMO_CONC_NM

    # Use realistic KD ~150 nM: kd/ka = 150e-9  → kd = 150e-9 * 1.5e5 = 0.0225 s⁻¹
    kd_per_s = ka_per_Ms * 150e-9  # forces KD = 150 nM

    rng = np.random.default_rng(seed)
    cycle_len_s = baseline_s + assoc_s + dissoc_s
    dt = 1.0 / rate_hz

    all_times = []
    wl_by_cycle: dict[int, dict[str, np.ndarray]] = {}
    cycle_windows: list[tuple[float, float]] = []

    t_offset = 0.0
    for ci, conc_nm in enumerate(concentrations_nm):
        conc_M = conc_nm * 1e-9
        kobs = ka_per_Ms * conc_M + kd_per_s   # pseudo-first-order rate

        times = np.arange(0, cycle_len_s, dt)
        t_abs = times + t_offset
        all_times.append(t_abs)

        assoc_start = baseline_s
        assoc_end = baseline_s + assoc_s
        dissoc_end = assoc_end + dissoc_s

        wl_cycle: dict[str, np.ndarray] = {}
        for ch, p in _CHANNEL_PROFILES.items():
            rmax_nm = p["rmax"]
            noise_nm = p["noise"]

            signal = np.full(len(times), _BASELINE_NM, dtype=float)

            # Association: Rmax × (1 - exp(-kobs × t))
            assoc_mask = (times >= assoc_start) & (times < assoc_end)
            t_a = times[assoc_mask] - assoc_start
            response_assoc = rmax_nm * (1 - np.exp(-kobs * t_a))
            signal[assoc_mask] += response_assoc

            assoc_end_val = float(response_assoc[-1]) if assoc_mask.any() else 0.0

            # Dissociation: level × exp(-kd × t)
            dissoc_mask = (times >= assoc_end) & (times < dissoc_end)
            t_d = times[dissoc_mask] - assoc_end
            signal[dissoc_mask] += assoc_end_val * np.exp(-kd_per_s * t_d)

            signal += rng.normal(0, noise_nm, len(times))
            wl_cycle[ch] = signal

        wl_by_cycle[ci] = wl_cycle
        cycle_windows.append((t_offset, t_offset + cycle_len_s))
        t_offset += cycle_len_s

    times_abs = np.concatenate(all_times)
    return times_abs, wl_by_cycle, cycle_windows


def load_demo_data_into_app(app) -> None:
    """Populate live graphs + Edits tab with 5-concentration kinetics SPR data.

    Live sensorgram: full 5-concentration series (~30 min, all cycles visible).
    Active cycle graph: zoomed to first cycle (0 → 360 s, full single injection).
    Edits tab: loaded from the matching Excel file via normal import path.

    Call from the Application instance (main.py) after hardware init.
    Wired to Ctrl+Shift+D.
    """
    import traceback as _tb
    _errors = []

    # ── Generate + load Excel into Edits tab (via main_window, not app) ───────
    try:
        excel_path = generate_demo_excel()
        app.main_window._load_data_from_excel_internal(excel_path)
    except Exception as _e:
        _errors.append(f"Excel load: {_e}\n{_tb.format_exc()}")

    # ── Generate concentration series for live buffer ─────────────────────────
    # Each cycle: 60s baseline + 120s assoc + 180s dissoc = 360s per cycle × 5 = 1800s total
    BASELINE_S = 60.0
    ASSOC_S    = 120.0
    DISSOC_S   = 180.0
    CYCLE_S    = BASELINE_S + ASSOC_S + DISSOC_S  # 360s per cycle
    times_abs, wl_by_cycle, cycle_windows = generate_concentration_series(
        concentrations_nm=_DEMO_CONC_NM,
        baseline_s=BASELINE_S,
        assoc_s=ASSOC_S,
        dissoc_s=DISSOC_S,
        rate_hz=2.0,   # 2 Hz for smooth live graph (vs 1 Hz for Excel)
        ka_per_Ms=5e4,
    )

    # Flatten per-cycle wl_by_cycle into single array per channel
    times_rel = times_abs - times_abs[0]   # start at t=0
    wl: dict[str, np.ndarray] = {}
    for ch in ("a", "b", "c", "d"):
        wl[ch] = np.concatenate([wl_by_cycle[ci][ch] for ci in range(len(wl_by_cycle))])

    epoch = time.time() - times_rel[-1]  # fake epoch so series ends "now"
    abs_times = times_rel + epoch

    # ── Live sensorgram buffers ───────────────────────────────────────────────
    try:
        buf = app.buffer_mgr
        app.clock.lock_display_offset(times_rel[1])
        display_offset = app.clock.display_offset

        for ch in ("a", "b", "c", "d"):
            buf.timeline_data[ch].time = times_rel.copy()
            buf.timeline_data[ch].timestamp = abs_times.copy()
            buf.timeline_data[ch].wavelength = wl[ch].copy()
            buf.baseline_wavelengths[ch] = float(wl[ch][0])

        for ch in ("a", "b", "c", "d"):
            app._pending_graph_updates[ch] = {"channel": ch}

        display_times = times_rel[1:] - display_offset
        for ch, idx in app._channel_pairs:
            curve = app.main_window.full_timeline_graph.curves[idx]
            curve.setData(display_times, wl[ch][1:])
    except Exception as _e:
        _errors.append(f"Live buffers: {_e}\n{_tb.format_exc()}")
        display_times = times_rel[1:]

    # Position cursors to show the LAST (highest concentration) full cycle
    # Last cycle starts at (N-1)*CYCLE_S, ends at N*CYCLE_S
    try:
        ftg = app.main_window.full_timeline_graph
        last_cycle_start = (len(_DEMO_CONC_NM) - 1) * CYCLE_S
        last_cycle_end   = len(_DEMO_CONC_NM) * CYCLE_S
        _dt = times_rel[1] - times_rel[0]  # time step (0.5s at 2Hz)
        _n_offset = int(last_cycle_start / _dt)
        _n_end    = int(last_cycle_end   / _dt) - 2

        # Clamp to valid range
        max_idx = len(display_times) - 1
        _n_offset = min(_n_offset, max_idx)
        _n_end    = min(_n_end,    max_idx)

        ftg.start_cursor.setValue(display_times[_n_offset])
        ftg.stop_cursor.setValue(display_times[_n_end])
    except Exception as _e:
        _errors.append(f"Cursors: {_e}")

    # Trigger cycle-of-interest update
    try:
        app._update_cycle_of_interest_graph()
    except Exception as _e:
        _errors.append(f"Cycle graph: {_e}")

    # ── Populate cycle queue table ────────────────────────────────────────────
    try:
        _populate_demo_cycle_queue(app)
    except Exception as _e:
        _errors.append(f"Cycle queue: {_e}\n{_tb.format_exc()}")

    # ── Sparq, spectra, edits table, bar chart ────────────────────────────────
    _populate_demo_extras(app)

    # ── Report any errors to log ──────────────────────────────────────────────
    if _errors:
        import logging as _logging
        _log = _logging.getLogger(__name__)
        for err in _errors:
            _log.error(f"[Demo] {err}")


def _populate_demo_cycle_queue(app) -> None:
    """Fill the QueueSummaryWidget with realistic demo Cycle objects."""
    from affilabs.domain.cycle import Cycle

    qp = getattr(app, 'queue_presenter', None)
    if qp is None:
        return

    # (type, length_minutes, name, status)
    _DEMO_CYCLES = [
        ("Baseline",       2.0,  "PBS baseline",       "completed"),
        ("Activation",     7.0,  "EDC/NHS",            "completed"),
        ("Immobilization", 30.0, "Anti-hIgG ligand",   "completed"),
        ("Blocking",       7.0,  "Ethanolamine",       "completed"),
        ("Baseline",       5.0,  "Stabilization",      "completed"),
        ("Binding",        8.5,  "hIgG 10 nM",         "running"),
        ("Binding",        8.5,  "hIgG 50 nM",         "pending"),
        ("Binding",        8.5,  "hIgG 100 nM",        "pending"),
        ("Binding",        8.5,  "hIgG 500 nM",        "pending"),
        ("Binding",        8.5,  "hIgG 1 µM",          "pending"),
    ]

    qp.clear_queue()
    for i, (ctype, dur, name, status) in enumerate(_DEMO_CYCLES):
        cycle = Cycle(
            type=ctype,
            length_minutes=dur,
            name=name,
            cycle_num=i + 1,
            status=status,
        )
        qp.add_cycle(cycle)

    # Refresh the visible widget
    try:
        app.main_window.sidebar.summary_table.refresh()
    except Exception:
        pass


# ── Excel demo file generator ─────────────────────────────────────────────────

#: Default output path for the kinetics demo Excel file
DEMO_EXCEL_PATH = os.path.join("_data", "demo", "kinetics_demo.xlsx")

# Concentrations for the 5 binding cycles (nM)
_DEMO_CONC_NM = [10.0, 50.0, 100.0, 500.0, 1000.0]

# RU per nm scale factor (1 nm shift ≈ 355 RU for gold SPR)
_NM_TO_RU = 355.0

# KD assumed for Langmuir isotherm in demo (nM)
_DEMO_KD_NM = 150.0


def generate_demo_excel(output_path: str = DEMO_EXCEL_PATH) -> str:
    """Generate a full kinetics demo .xlsx file loadable by the Edits tab.

    Creates four sheets:
      - Channel Data  : side-by-side time+wavelength for A/B/C/D
      - Cycles        : 10 cycle rows (Baseline→Binding series) with ΔSPR
      - Metadata      : experiment info key/value pairs
      - Summary       : human-readable overview for screenshots

    Returns the absolute path of the written file.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # Generate 5 distinct concentration-series sensorgrams
    # Each cycle: 60s baseline + 120s association + 180s dissociation = 360s
    # Short association window ensures curves don't all plateau — concentration effect is visible
    BASELINE_S = 60.0
    ASSOC_S = 120.0
    DISSOC_S = 180.0
    KA_PER_MS = 5e4   # slower ka so kobs varies meaningfully across [C] range
    times_abs, wl_by_cycle, cycle_windows = generate_concentration_series(
        concentrations_nm=_DEMO_CONC_NM,
        baseline_s=BASELINE_S,
        assoc_s=ASSOC_S,
        dissoc_s=DISSOC_S,
        rate_hz=1.0,
        ka_per_Ms=KA_PER_MS,
    )

    # Compute ΔSPR at end-of-association for each cycle (plateau value)
    ka_per_Ms = KA_PER_MS
    kd_per_s = ka_per_Ms * 150e-9
    rmax_ru = {ch: _CHANNEL_PROFILES[ch]["rmax"] * _NM_TO_RU for ch in "abcd"}

    wb = openpyxl.Workbook()

    # ── Sheet 1: Raw Data (flat, interleaved — read by loader format 1) ────────
    ws_raw = wb.active
    ws_raw.title = "Raw Data"

    raw_hdr = ["time", "channel", "value"]
    for col_idx, h in enumerate(raw_hdr, 1):
        cell = ws_raw.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1D6FA5")
        cell.alignment = Alignment(horizontal="center")

    raw_row = 2
    n_total = len(times_abs)
    # times_abs is already 1 Hz; write all 4 channels interleaved per timestep
    # We need to reconstruct per-timestep from the by-cycle structure
    for ci, (t_start_c, t_end_c) in enumerate(cycle_windows):
        wl_c = wl_by_cycle[ci]
        cycle_times = times_abs[(times_abs >= t_start_c) & (times_abs < t_end_c)]
        n_c = len(wl_c["a"])
        for j in range(n_c):
            t_val = round(float(t_start_c) + j, 2)
            for ch in ("a", "b", "c", "d"):
                ws_raw.cell(row=raw_row, column=1, value=t_val)
                ws_raw.cell(row=raw_row, column=2, value=ch)
                ws_raw.cell(row=raw_row, column=3, value=round(float(wl_c[ch][j]), 4))
                raw_row += 1

    for col_idx, w in enumerate([12, 10, 14], 1):
        ws_raw.column_dimensions[get_column_letter(col_idx)].width = w

    # ── Sheet 2: Channel Data (side-by-side — for viewing in Excel) ───────────
    ws_ch = wb.create_sheet("Channel Data")

    hdr = []
    for ch in ("A", "B", "C", "D"):
        hdr += [f"Time {ch} (s)", f"Channel {ch} (nm)"]
    for col_idx, h in enumerate(hdr, 1):
        cell = ws_ch.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1D6FA5")
        cell.alignment = Alignment(horizontal="center")

    ch_row = 2
    for ci, (t_start_c, _) in enumerate(cycle_windows):
        wl_c = wl_by_cycle[ci]
        n_c = len(wl_c["a"])
        for j in range(n_c):
            t_val = round(float(t_start_c) + j, 2)
            for ch_col, ch in enumerate(("a", "b", "c", "d")):
                base_col = ch_col * 2 + 1
                ws_ch.cell(row=ch_row, column=base_col, value=t_val)
                ws_ch.cell(row=ch_row, column=base_col + 1, value=round(float(wl_c[ch][j]), 4))
            ch_row += 1

    for col_idx in range(1, 9):
        ws_ch.column_dimensions[get_column_letter(col_idx)].width = 16

    # ── Sheet 3: Cycles ───────────────────────────────────────────────────────
    ws_cy = wb.create_sheet("Cycles")

    cycle_cols = [
        "cycle_num", "type", "name",
        "start_time_sensorgram", "end_time_sensorgram",
        "duration_minutes", "concentration_value",
        "delta_ch1", "delta_ch2", "delta_ch3", "delta_ch4",
        "delta_measured", "delta_ref_ch", "note", "flags",
    ]
    for col_idx, h in enumerate(cycle_cols, 1):
        cell = ws_cy.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1D6FA5")
        cell.alignment = Alignment(horizontal="center")

    names = ["hIgG 10 nM", "hIgG 50 nM", "hIgG 100 nM", "hIgG 500 nM", "hIgG 1 µM"]

    cycle_row = 2
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Baseline row — uses first 60s of first cycle
    t0 = cycle_windows[0][0]
    baseline_row_data = [1, "Baseline", "PBS baseline",
                         round(t0 + 1, 1), round(t0 + BASELINE_S, 1),
                         round(BASELINE_S / 60, 1),
                         "", "", "", "", "", False, "None", "", ""]
    for col_idx, val in enumerate(baseline_row_data, 1):
        cell = ws_cy.cell(row=cycle_row, column=col_idx, value=val)
        cell.border = border
        if cycle_row % 2 == 0:
            cell.fill = PatternFill("solid", fgColor="F5F9FC")
    cycle_row += 1

    # Binding cycles — each with its own time window and concentration-scaled ΔSPR
    for i, (conc, name, (t_start_c, t_end_c)) in enumerate(
        zip(_DEMO_CONC_NM, names, cycle_windows)
    ):
        conc_M = conc * 1e-9
        kobs = ka_per_Ms * conc_M + kd_per_s
        # ΔSPR = Rmax × (1 - exp(-kobs × assoc_duration)) × NM_TO_RU
        frac = 1.0 - np.exp(-kobs * ASSOC_S)
        assoc_start_abs = t_start_c + BASELINE_S
        assoc_end_abs = assoc_start_abs + ASSOC_S
        dur_b = round((assoc_end_abs - assoc_start_abs + DISSOC_S) / 60.0, 1)
        row_data = [
            2 + i, "Binding", name,
            round(assoc_start_abs, 1), round(t_end_c, 1), dur_b,
            str(conc),
            round(rmax_ru["a"] * frac, 1),
            round(rmax_ru["b"] * frac, 1),
            round(rmax_ru["c"] * frac, 1),
            round(rmax_ru["d"] * frac, 1),
            True, "None", "", "",
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_cy.cell(row=cycle_row, column=col_idx, value=val)
            cell.border = border
            if cycle_row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F5F9FC")
        cycle_row += 1

    # Column widths for Cycles sheet
    col_widths = [10, 16, 22, 22, 22, 16, 18, 12, 12, 12, 12, 14, 14, 14, 10]
    for col_idx, w in enumerate(col_widths, 1):
        ws_cy.column_dimensions[get_column_letter(col_idx)].width = w

    # ── Sheet 3: Metadata ─────────────────────────────────────────────────────
    ws_meta = wb.create_sheet("Metadata")
    ws_meta.cell(row=1, column=1, value="key").font = Font(bold=True)
    ws_meta.cell(row=1, column=2, value="value").font = Font(bold=True)
    metadata_rows = [
        ("experiment_name",  "hIgG Dose-Response Kinetics"),
        ("analyte",          "Human IgG (hIgG)"),
        ("ligand",           "Anti-hIgG antibody"),
        ("chip_chemistry",   "EDC/NHS amine coupling"),
        ("buffer",           "PBS pH 7.4"),
        ("flow_rate",        "25 µL/min"),
        ("temperature",      "25°C"),
        ("instrument",       "Affilabs.core P4SPR"),
        ("operator",         "Demo"),
        ("date",             time.strftime("%Y-%m-%d")),
        ("KD_estimated_nM",  str(_DEMO_KD_NM)),
        ("notes",            "Kinetics demo dataset — generated by Affilabs.core"),
    ]
    for row_idx, (k, v) in enumerate(metadata_rows, 2):
        ws_meta.cell(row=row_idx, column=1, value=k)
        ws_meta.cell(row=row_idx, column=2, value=v)
    ws_meta.column_dimensions["A"].width = 22
    ws_meta.column_dimensions["B"].width = 42

    # ── Sheet 4: Summary ──────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")
    summary_lines = [
        ("hIgG Dose-Response — Kinetics Demo", True),
        ("", False),
        ("4 SPR channels (A–D) measuring hIgG binding to immobilised anti-hIgG.", False),
        ("Experiment: Baseline → Activation → Immobilization → Blocking → 5× Binding", False),
        ("", False),
        ("Channel  Rmax (RU)  KD est.  Profile", True),
        (f"A        {rmax_ru['a']:.0f}        150 nM   Fast on / slow off", False),
        (f"B        {rmax_ru['b']:.0f}        150 nM   Slow on / fast off", False),
        (f"C        {rmax_ru['c']:.0f}        150 nM   Fast on / medium off", False),
        (f"D        {rmax_ru['d']:.0f}        150 nM   Very slow on/off", False),
        ("", False),
        ("Concentrations: 10, 50, 100, 500, 1000 nM", False),
        ("Generated by Affilabs.core demo mode (Ctrl+Shift+D)", False),
    ]
    for row_idx, (text, bold) in enumerate(summary_lines, 1):
        cell = ws_sum.cell(row=row_idx, column=1, value=text)
        if bold:
            cell.font = Font(bold=True, size=12)
    ws_sum.column_dimensions["A"].width = 60

    abs_path = os.path.abspath(output_path)
    wb.save(abs_path)
    return abs_path


# ── Demo extras: Sparq, spectra, edits table, binding chart ──────────────────


def _make_transmission_spectrum(
    baseline_nm: float,
    shift_nm: float,
    rng: np.random.Generator,
) -> tuple[np.ndarray, np.ndarray]:
    """Generate a realistic SPR transmission spectrum with a dip.

    Returns (wavelengths, transmission) over 560–720 nm.
    Shape: nearly flat broadband envelope (~75%) with a clear SPR dip
    (inverted absorption band) cutting down to ~20% at the resonance.
    SPR dip position = baseline_nm + shift_nm (red shift on binding).
    """
    wl = np.linspace(560, 720, 512)

    # Broad, nearly flat LED envelope — wide Gaussian so edges roll off gently
    # but the active window 580–700 nm stays nearly flat (~75%)
    envelope = 75.0 * np.exp(-0.5 * ((wl - 635) / 75) ** 2)

    # SPR dip: Lorentzian, FWHM ~30 nm, cuts down to ~25% of local envelope
    dip_centre = baseline_nm + shift_nm
    dip_depth = 0.72   # fraction removed at dip centre
    dip_hwhm = 15.0
    dip_factor = 1.0 - dip_depth / (1.0 + ((wl - dip_centre) / dip_hwhm) ** 2)

    transmission = envelope * dip_factor

    # Mild noise
    transmission += rng.normal(0, 0.25, len(wl))
    transmission = np.clip(transmission, 0, 100)

    return wl, transmission


def _populate_demo_extras(app) -> None:
    """Populate Sparq message, transmission spectra, edits table, and bar chart.

    Called from load_demo_data_into_app() after sensorgrams are loaded.
    All exceptions are swallowed individually so partial failure doesn't abort.
    """
    rng = np.random.default_rng(42)

    # ── 1. Sparq welcome message ──────────────────────────────────────────────
    try:
        sidebar = getattr(app.main_window, 'spark_sidebar', None)
        spark = getattr(sidebar, 'spark_widget', None) if sidebar else None
        if spark is None:
            spark = getattr(app.main_window, 'spark_widget', None)
        if spark is not None:
            spark.push_system_message(
                "Demo mode active — I'm analysing your hIgG dose-response experiment.\n\n"
                "**Channels A–D** show 4 independent SPR sensors with distinct on/off "
                "kinetics. Channel A has the strongest binding (Rmax ≈ 994 RU), "
                "Channel D the weakest (Rmax ≈ 320 RU).\n\n"
                "The current cycle is **hIgG 10 nM**. Signal quality is excellent on "
                "all channels (IQ: ●●●●). Try the **Edits** tab to view the full "
                "concentration series and generate a binding plot."
            )
    except Exception:
        pass

    # ── 2. Fake transmission spectra ─────────────────────────────────────────
    try:
        sp = getattr(app, 'spectroscopy_presenter', None)
        if sp is None:
            ui_coord = getattr(app, 'ui_updates', None)
            sp = getattr(ui_coord, 'spectroscopy_presenter', None)
        if sp is not None:
            # Representative end-of-association shift for each channel
            shifts = {"a": 2.8, "b": 1.6, "c": 2.1, "d": 0.9}
            for ch, shift in shifts.items():
                wl, tr = _make_transmission_spectrum(_BASELINE_NM, shift, rng)
                sp.update_transmission(ch, wl, tr)
    except Exception:
        pass

    # ── 3. Edits table: populated via Excel load (generate_demo_excel) ───────
    # No add_cycle() calls here — the Excel import already populates the table.

    # ── 4. Delta SPR bar chart: driven by cursor placement, not pre-populated ─
    # The bar chart updates correctly when the user selects a cycle and places
    # the ΔSPR cursors. Pre-populating with hardcoded values causes conflicts.


# ── Legacy function kept for compatibility ────────────────────────────────────

def generate_demo_spr_data(
    duration_seconds: float = 600,
    sampling_rate: float = 2.0,
    num_channels: int = 4,
    baseline_ru: float = 0.0,
    max_response_ru: dict[str, float] | None = None,
    association_start: float = 120,
    association_duration: float = 240,
    dissociation_start: float = 360,
    ka: float = 1e5,
    kd: float = 1e-3,
    noise_level: float = 0.5,
    seed: int | None = 42,
) -> tuple[np.ndarray, dict[str, np.ndarray]]:
    """Legacy RU-domain demo generator. Use generate_demo_sensorgrams() instead."""
    if seed is not None:
        np.random.seed(seed)

    num_points = int(duration_seconds * sampling_rate)
    time_array = np.linspace(0, duration_seconds, num_points)

    if max_response_ru is None:
        max_response_ru = {"a": 55.0, "b": 48.0, "c": 52.0, "d": 45.0}

    channel_names = ["a", "b", "c", "d"][:num_channels]
    channel_data = {}
    C = 10e-9
    kobs = ka * C + kd

    for ch in channel_names:
        rmax = max_response_ru.get(ch, 50.0)
        signal = baseline_ru + np.linspace(0, 0.5, num_points)

        assoc_mask = (time_array >= association_start) & (time_array < dissociation_start)
        t_a = time_array[assoc_mask] - association_start
        signal[assoc_mask] += rmax * (1 - np.exp(-kobs * t_a))

        steady = rmax * (1 - np.exp(-kobs * (dissociation_start - association_start)))
        dissoc_mask = time_array >= dissociation_start
        t_d = time_array[dissoc_mask] - dissociation_start
        signal[dissoc_mask] += steady * np.exp(-kd * t_d)

        signal += np.random.normal(0, noise_level, num_points)
        signal += 0.2 * np.sin(2 * np.pi * 0.01 * time_array)
        channel_data[ch] = signal

    return time_array, channel_data
